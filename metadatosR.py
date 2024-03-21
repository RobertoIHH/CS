import os
import sys
import fitz
from pathlib import Path
from openpyxl import load_workbook
from docx import Document

def obtener_metadatos_docx(docx_path):
    try:
        document = Document(docx_path)
        metadata = {}
        core_props = document.core_properties
        # Verificar si el atributo 'title' está presente en core_properties
        if hasattr(core_props, 'title'):
            metadata['Título'] = core_props.title
        # Verificar si el atributo 'author' está presente en core_properties
        if hasattr(core_props, 'author'):
            metadata['Autor'] = core_props.author
        # Verificar si el atributo 'subject' está presente en core_properties
        if hasattr(core_props, 'subject'):
            metadata['Asunto'] = core_props.subject
        # Verificar si el atributo 'creator' está presente en core_properties
        if hasattr(core_props, 'creator'):
            metadata['Creador'] = core_props.creator
        # Verificar si el atributo 'keywords' está presente en core_properties
        if hasattr(core_props, 'keywords'):
            metadata['Palabras clave'] = core_props.keywords
        # Verificar si el atributo 'created' está presente en core_properties
        if hasattr(core_props, 'created'):
            metadata['Fecha de creación'] = core_props.created
        # Verificar si el atributo 'last_printed' está presente en core_properties
        if hasattr(core_props, 'last_printed'):
            metadata['Última impresión'] = core_props.last_printed
        # Verificar si el atributo 'modified' está presente en core_properties
        if hasattr(core_props, 'modified'):
            metadata['Última modificación'] = core_props.modified
        return metadata
    except Exception as e:
        print(f"Error al obtener metadatos de {docx_path}: {e}")
        return None

def obtener_metadatos_xlsx(xlsx_path):
    try:
        metadata = {}
        workbook = load_workbook(filename=xlsx_path, read_only=True)
        workbook_properties = workbook.properties
        # Obtener los metadatos del documento XLSX
        metadata['libro']= {}
        metadata['libro']['Título'] = workbook_properties.title
        metadata['libro']['Autor'] = workbook_properties.creator
        metadata['libro']['Fecha de creación'] = workbook_properties.created
        metadata['libro']['Última modificación'] = workbook_properties.modified
        metadata['libro']['Número de hojas'] = len(workbook.sheetnames)
        
        # Propiedades de las hojas
        sheets_info = {}
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet_info = {
                'Nombre': sheet.title,
                'Filas': sheet.max_row,
                'Columnas': sheet.max_column
            }
            sheets_info[sheet_name] = sheet_info
        metadata['Hojas'] = sheets_info
        return metadata
    except Exception as e:
        print(f"Error al obtener metadatos de {xlsx_path}: {e}")
        return None

def obtener_metadatos_pdf(pdf_path):
    try:
        metadata = {}
        with fitz.open(pdf_path) as pdf_document:
            # Obtener los metadatos del documento PDF
            metadata = pdf_document.metadata
        return metadata
    except Exception as e:
        print(f"Error al obtener metadatos de {pdf_path}: {e}")
        return None

def buscar_archivos_ext(carpeta, extension):
    archivos = []
    # Recorrer la carpeta y sus subcarpetas en busca de archivos con la extensión proporcionada
    for root, dirs, files in os.walk(carpeta):
        for file in files:
            if file.endswith(extension):
                archivos.append(os.path.join(root, file))  # Agregar la ruta del archivo a la lista
    return archivos

def obtener_metadatos_archivos(carpeta, extension):
    archivos = buscar_archivos_ext(carpeta, extension)  # Obtener archivos con la extensión proporcionada
    metadatos = []
    # Obtener los metadatos de cada archivo según su extensión
    for archivo in archivos:
        if extension == '.docx':
            meta = obtener_metadatos_docx(archivo)
        elif extension == '.xlsx':
            meta = obtener_metadatos_xlsx(archivo)
        elif extension == '.pdf':
            meta = obtener_metadatos_pdf(archivo)
        if meta:
            metadatos.append((archivo, meta))  # Agregar los metadatos a la lista
    return metadatos

# Ruta a la carpeta que contiene archivos
carpeta = input("Ingrese la ruta de los archivos ejemplo(/home/zicfro/Downloads): ")
if not os.path.isdir(carpeta):
       print("La ruta ingresada no es válida.")
       sys.exit();
# Solicitar al usuario la extensión de los archivos
extension = input("Ingrese la extensión de los archivos (.docx, .xlsx o .pdf): ").lower()

# Obtener y mostrar los metadatos de los archivos según la extensión
if extension == '.docx':
    print("Metadatos de los archivos DOCX encontrados:")
elif extension == '.xlsx':
    print("Metadatos de los archivos XLSX encontrados:")
elif extension == '.pdf':
    print("Metadatos de los archivos PDF encontrados:")
else:
    print("Extensión no válida.")
    exit()

# Obtener y mostrar los metadatos de los archivos encontrados
metadatos_archivos = obtener_metadatos_archivos(carpeta, extension)
if metadatos_archivos:
    for archivo, metadata in metadatos_archivos:
        file_name = Path(archivo).stem
        print(f"\nArchivo: {file_name}{extension}")
        print("Metadatos:")
        if extension == '.xlsx':
            print("Libro:")
            for clave, valor in metadata.items():
                if isinstance(valor, dict):
                    print(f"\t{clave}:")
                    for subclave, subvalor in valor.items():
                        print(f"\t\t{subclave}: {subvalor}")
                else:
                    print(f"\t{clave}: {valor}")
                    
        else:
            for clave, valor in metadata.items():
                print(f"\t{clave}: {valor}")
else:
    print(f"No se encontraron archivos {extension.upper()} en la carpeta {carpeta}.")
